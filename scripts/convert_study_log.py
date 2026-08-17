#!/usr/bin/env python3
"""
Convert a study-log Excel workbook into the JSON format consumed by the
"Import Incremental Rems with History" command of the Incremental RemNote
RemNote plugin.

NOTE: this script is a SAMPLE, tailored to one specific spreadsheet layout
(including source-specific adjustments like FLASHCARD_TIME_FACTOR below).
Any tool that produces the version-1 JSON format documented in the wiki
(Plugin Commands Reference → Import Incremental Rems with History) works.

Expected workbook layout (matching "Log de Atividades"):
  - Sheet "Literatura": Item (A), Livro (E)
  - Sheet "Log": DATE (A), TMP PONDERADO (I), CICLO (K), INICIAL (M),
    FINAL (N), PÁGINAS ou CARDS (O), CÓD MAT (S), CHAPTERS (U), OBS (V)

Output structure (version 1):
{
  "version": 1,
  "generatedAt": "...",
  "defaultPriority": 90,
  "nextRepDays": 10,
  "books": [
    {
      "item": "2.01",
      "title": "2.01 - Arte Naval ...",
      "history": [IncrementalRep, ...] | null,   // logs with empty chapter
      "chapters": [
        { "chapter": "1", "title": "Chapter 1", "history": [IncrementalRep, ...] }
      ]
    }
  ]
}

Each history is sorted by date and each log becomes an eventType 'rep' entry
with reviewTimeSeconds (weighted hours * 3600), interval (days until next rep)
and notes assembled from CICLO / pages / OBS. The plugin appends a
'madeIncremental' marker (stamped with nextRepMs) at import time, so the
scheduler restarts interval counting from the import.

Usage:
  python3 scripts/convert_study_log.py <input.xlsm> [output.json] [--before YYYY-MM-DD]

Options:
  --before YYYY-MM-DD   Only include log rows dated strictly BEFORE this date
                        (e.g. to exclude sessions already tracked natively).
"""

import datetime
import json
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip3 install --user openpyxl")

DEFAULT_PRIORITY = 90
NEXT_REP_DAYS = 10

# Sessions whose time is mostly already accounted for inside RemNote's own
# flashcard review stats — only the missing fraction of their logged time is
# imported, to avoid double counting.
DISCOUNTED_CICLOS = {'RemNote Queue', 'Anki Cards', 'Revisão'}
FLASHCARD_TIME_FACTOR = 0.29

# Log sheet column indexes (0-based)
COL_DATE = 0        # A
COL_TMP_POND = 8    # I  TMP PONDERADO (weighted hours)
COL_CICLO = 10      # K
COL_INICIAL = 12    # M
COL_FINAL = 13      # N
COL_PAGINAS = 14    # O  PÁGINAS ou CARDS
COL_COD = 18        # S  CÓD MAT
COL_CHAPTERS = 20   # U
COL_OBS = 21        # V

# Literatura sheet column indexes (0-based)
LIT_COL_ITEM = 0    # A
LIT_COL_LIVRO = 4   # E


def clean(value):
    """Normalize a cell to a stripped single-line string, or None."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return re.sub(r'\s*[\r\n]+\s*', '; ', s)


def fmt_num(value):
    """Render 44.0 as '44' but keep genuine decimals."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def chapter_title(chapter: str) -> str:
    """'1' -> 'Chapter 1'; 'Introdução' stays as-is."""
    # Chapter-number-ish values: digits, roman numerals, appendix letters,
    # possibly in comma/'e'-separated lists. Word-like values (e.g.
    # 'Introdução') are kept verbatim.
    if re.fullmatch(r'[0-9IVXA-Z][0-9IVXA-Z\s,.eE&\-]*', chapter):
        return f'Chapter {chapter}'
    return chapter


def to_ms(dt: datetime.datetime) -> int:
    return int(dt.timestamp() * 1000)


def build_notes(row) -> str | None:
    parts = []
    ciclo = clean(row[COL_CICLO])
    if ciclo:
        parts.append(f'Ciclo: {ciclo}')

    inicial = row[COL_INICIAL]
    final = row[COL_FINAL]
    paginas = row[COL_PAGINAS]
    if inicial is not None and final is not None:
        pages = f'{fmt_num(inicial)}–{fmt_num(final)}'
        if paginas is not None:
            pages += f' ({fmt_num(paginas)})'
        parts.append(f'Páginas: {pages}')
    elif paginas is not None:
        parts.append(f'Páginas: {fmt_num(paginas)}')

    obs = clean(row[COL_OBS])
    if obs:
        parts.append(f'OBS: {obs}')
    return ' | '.join(parts) if parts else None


def build_history(rows) -> list[dict]:
    """rows: log rows of one (item, chapter) group -> IncrementalRep[]"""
    rows = sorted(rows, key=lambda r: r[COL_DATE])
    history = []

    for i, row in enumerate(rows):
        ms = to_ms(row[COL_DATE])
        entry = {
            'date': ms,
            'scheduled': ms,
            'eventType': 'rep',
        }
        tmp = row[COL_TMP_POND]
        if isinstance(tmp, (int, float)) and tmp > 0:
            ciclo = clean(row[COL_CICLO])
            factor = FLASHCARD_TIME_FACTOR if ciclo in DISCOUNTED_CICLOS else 1.0
            entry['reviewTimeSeconds'] = round(tmp * factor * 3600)
        if i + 1 < len(rows):
            gap_days = (rows[i + 1][COL_DATE] - row[COL_DATE]).days
            entry['interval'] = max(gap_days, 0)
        notes = build_notes(row)
        if notes:
            entry['notes'] = notes
        history.append(entry)

    return history


def main():
    args = sys.argv[1:]
    before = None
    if '--before' in args:
        i = args.index('--before')
        try:
            before = datetime.datetime.strptime(args[i + 1], '%Y-%m-%d')
        except (IndexError, ValueError):
            sys.exit('--before requires a date in YYYY-MM-DD format')
        del args[i:i + 2]
    if not args:
        sys.exit(__doc__)
    input_path = args[0]
    output_path = args[1] if len(args) > 1 else 'study-log-import.json'

    print(f'Reading {input_path} ...')
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    livros = {}
    for row in wb['Literatura'].iter_rows(min_row=2, values_only=True):
        item = clean(row[LIT_COL_ITEM])
        if item:
            livros[item] = clean(row[LIT_COL_LIVRO]) or '(sem título)'

    groups = defaultdict(list)   # (item, chapter) -> [rows]
    skipped = []
    excluded_by_date = 0
    for idx, row in enumerate(wb['Log'].iter_rows(min_row=2, values_only=True), start=2):
        if row[COL_DATE] is None or not isinstance(row[COL_DATE], datetime.datetime):
            continue
        if before is not None and row[COL_DATE] >= before:
            excluded_by_date += 1
            continue
        cod = clean(row[COL_COD])
        if not cod:
            skipped.append((idx, 'empty CÓD MAT'))
            continue
        if cod not in livros:
            skipped.append((idx, f'CÓD MAT {cod!r} not in Literatura'))
            continue
        chapter = clean(row[COL_CHAPTERS]) or ''
        groups[(cod, chapter)].append(row)

    books = []
    def item_sort_key(s):
        return [(0, int(p)) if p.isdigit() else (1, p) for p in s.split('.')]

    for item in sorted({cod for cod, _ in groups}, key=item_sort_key):
        title = f'{item} - {livros[item]}'
        book = {'item': item, 'title': title, 'history': None, 'chapters': []}
        chapters = sorted(
            (ch for cod, ch in groups if cod == item),
            key=lambda c: (c != '', not c[:1].isdigit(), len(c), c),
        )
        for ch in chapters:
            hist = build_history(groups[(item, ch)])
            if ch == '':
                book['history'] = hist
            else:
                book['chapters'].append({
                    'chapter': ch,
                    'title': chapter_title(ch),
                    'history': hist,
                })
        books.append(book)

    payload = {
        'version': 1,
        'generatedAt': datetime.datetime.now().isoformat(timespec='seconds'),
        'defaultPriority': DEFAULT_PRIORITY,
        'nextRepDays': NEXT_REP_DAYS,
        'books': books,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    # ---- Report ----
    n_chapter_rems = sum(len(b['chapters']) for b in books)
    n_book_rems = sum(1 for b in books if b['history'])
    n_logs = sum(len(rs) for rs in groups.values())
    print(f'Books: {len(books)}  |  Incremental Rems: {n_book_rems + n_chapter_rems} '
          f'({n_book_rems} book-level + {n_chapter_rems} chapter-level)  |  Logs: {n_logs}')
    if before is not None:
        print(f'Excluded {excluded_by_date} log rows dated on/after {before.date()}')

    big = []
    for b in books:
        for name, hist in [(b['title'], b['history'])] + [
            (f"{b['title']} / {c['title']}", c['history']) for c in b['chapters']
        ]:
            if hist:
                size = len(json.dumps(hist, ensure_ascii=False).encode())
                if size > 50_000:
                    big.append((size, name, len(hist)))
    if big:
        print(f'\n⚠ Histories over 50 KB (verify these rems sync after import):')
        for size, name, n in sorted(big, reverse=True):
            print(f'  {size/1024:.0f} KB  {n:4d} entries  {name}')

    if skipped:
        print(f'\nSkipped {len(skipped)} log rows:')
        for row_idx, reason in skipped[:20]:
            print(f'  Log row {row_idx}: {reason}')

    print(f'\nWrote {output_path}')


if __name__ == '__main__':
    main()
